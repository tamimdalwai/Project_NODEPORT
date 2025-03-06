import asyncio
import sys
import logging
import os
import pandas as pd
from pymodbus.exceptions import ConnectionException, ModbusException
from pymodbus.client import AsyncModbusTcpClient
import time
from openpyxl import Workbook, load_workbook
from datetime import datetime
from threading import Lock

# Add global lock for data synchronization
data_lock = Lock()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("modbus_client.log"),
        logging.StreamHandler(sys.stdout)
    ]
)

# Create logger instances
logger = logging.getLogger("ModbusClient")
logging.getLogger("pymodbus").setLevel(logging.WARNING)

# Global dictionary to store latest Modbus data for all PLCs
latest_modbus_data = {}
plc_state = {}

# Configuration paths
SAVED_ADDRESS_FILE_PATH = "config/saveAddress.xlsx"
SELECTED_PLC_FILE = "config/selectedPlc.txt"


def get_modbus_addresses_with_check(sheet_name):
    """Read Modbus addresses from Excel sheet."""
    try:
        df = pd.read_excel(SAVED_ADDRESS_FILE_PATH, sheet_name=sheet_name)
    except Exception as e:
        logger.error(f"Error reading Excel file: {e}")
        return {"Coils": [], "Input Bits": [], "Analog Inputs": []}

    if df.empty:
        logger.warning(f"Sheet '{sheet_name}' is empty")
        return {"Coils": [], "Input Bits": [], "Analog Inputs": []}

    # Extract and adjust addresses
    coils = df['MODBUS ADDRESS (Coils)'].dropna().astype(int).tolist()
    input_bits = [addr - 10000 for addr in
                  df['MODBUS ADDRESS (Input Bits)'].dropna().astype(int).tolist()]
    analog_inputs = [addr - 30000 for addr in
                     df['MODBUS ADDRESS (Analog Inputs)'].dropna().astype(int).tolist()]

    return {
        "Coils": coils,
        "Input Bits": input_bits,
        "Analog Inputs": analog_inputs
    }


def get_coils(sheet_name):
    """Get coil addresses for selected PLC."""
    coils = get_modbus_addresses_with_check(sheet_name)["Coils"]
    logger.info(f"COILS: {coils}")
    return coils


def get_input_bits(sheet_name):
    """Get input bit addresses for selected PLC."""
    input_bits = get_modbus_addresses_with_check(sheet_name)["Input Bits"]
    logger.info(f"INPUT BITS: {input_bits}")
    return input_bits


def get_inputs_register(sheet_name):
    """Get input register addresses for selected PLC."""
    analog_inputs = get_modbus_addresses_with_check(sheet_name)["Analog Inputs"]
    logger.info(f"INPUT REGISTERS: {analog_inputs}")
    return analog_inputs

# field_data = {}
# def get_data_for_Plc(sheet_name):
#      field_data[plc_id] = {
#         "coil_states": get_coils(sheet_name),
#         "input_status_states": get_input_bits(sheet_name),
#         "input_register_states": get_inputs_register(sheet_name)
#      }
#     return field_data



async def check_connection(client, plc_id, retry_interval=5):
    """Verify Modbus connection with retry logic."""
    try:
        logger.info(f"Checking connection for {plc_id}...")
        connection_ok = False

        # Try reading different register types
        for read_func, addresses in [
            (client.read_coils, get_coils(plc_id)),
            (client.read_discrete_inputs, get_input_bits(plc_id)),
            (client.read_input_registers, get_inputs_register(plc_id))
        ]:
            if addresses:
                logger.debug(f"Check Connection {plc_id}: Addresses: {addresses}")
                response = await read_func(address=0, count=len(addresses))
                if not response.isError():
                    connection_ok = True
                    break

        plc_state[plc_id] = {"connected": connection_ok}
        logger.info(f"PLC {plc_id} connected: {connection_ok}")
        return connection_ok

    except Exception as e:
        logger.error(f"Inside check Connection error for {plc_id}: {e}")
        plc_state[plc_id] = {"connected": False}
        return False


async def append_to_excel(data):
    """Append Modbus data to separate Excel files based on PLC ID and register type."""
    plc_id = data["plc_id"]
    # coil_states = data["coil_states"]
    # input_status_states = data["input_status_states"]
    # Convert boolean values to integers (1 or 0)
    coil_states = {k: int(v) for k, v in data["coil_states"].items()}
    input_status_states = {k: int(v) for k, v in data["input_status_states"].items()}
    input_register_states = data["input_register_states"]

    folder_path = os.path.join("modbus_data", plc_id)
    os.makedirs(folder_path, exist_ok=True)

    try:
        await asyncio.gather(
            _append_register_data(os.path.join(folder_path, "coil_states.xlsx"),
                                  "Coil States", coil_states),
            _append_register_data(os.path.join(folder_path, "input_status_states.xlsx"),
                                  "Input Status States", input_status_states),
            _append_register_data(os.path.join(folder_path, "input_register_states.xlsx"),
                                  "Input Register States", input_register_states, True)
        )
    except Exception as e:
        logger.error(f"Excel append error for {plc_id}: {e}")


async def _append_register_data(file_path, sheet_name, data, is_input_register=False):
    """Helper function to append data to a specific Excel file."""
    if not data:
        return

    try:
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            addresses = [f"{addr:05}" for addr in sorted(data.keys())]
            ws.append(["Date", "Time"] + addresses)

        timestamp = datetime.now()
        values = [data[addr] for addr in sorted(data.keys())]
        ws.append([
                      timestamp.strftime("%Y-%m-%d"),
                      timestamp.strftime("%H:%M:%S")
                  ] + values)

        wb.save(file_path)
        logger.debug(f"Data saved to {file_path}")

    except Exception as e:
        logger.error(f"Error saving to {file_path}: {e}")


async def read_registers(client, read_func, addresses, sampling_frequency, max_count=2000):
    """Generic register reading function with address grouping."""
    results = {}
    try:
        sorted_addrs = sorted(addresses)
        start = sorted_addrs[0]
        prev_addr = start

        for addr in sorted_addrs[1:]:
            if addr != prev_addr + 1 or (addr - start + 1) > max_count:
                count = prev_addr - start + 1
                response = await read_func(address=start - 1, count=count)

                if not response.isError():
                    for i, a in enumerate(range(start, prev_addr + 1)):
                        if a in addresses:
                            results[a] = response.registers[i] if read_func == client.read_input_registers else \
                            response.bits[i]
                start = addr
            prev_addr = addr

        # Read last range
        count = prev_addr - start + 1
        response = await read_func(address=start - 1, count=count)

        if not response.isError():
            for i, a in enumerate(range(start, prev_addr + 1)):
                if a in addresses:
                    results[a] = response.registers[i] if read_func == client.read_input_registers else response.bits[i]

        await asyncio.sleep(sampling_frequency / 1000)
        return results

    except ModbusException as e:
        logger.error(f"Inside read register Modbus error: {e}")
        return {}
    except Exception as e:
        logger.error(f"Unexpected error: {e}")
        return {}


async def modbus_client_loop(plc_id, ip, port, sampling_frequency):
    """Main Modbus client loop for individual PLC."""
    logger.info(f"Starting Modbus client for {plc_id} at {ip}:{port}")

    while True:
        try:
            async with AsyncModbusTcpClient(ip, port=port) as client:
                logger.info(f"Connected to {plc_id} at {ip}:{port}")

                while True:
                    if not await check_connection(client, plc_id):
                        await asyncio.sleep(5)
                        continue

                    try:
                        # Read all registers in parallel
                        coil_task = read_registers(client, client.read_coils,
                                                   get_coils(plc_id), sampling_frequency)
                        input_task = read_registers(client, client.read_discrete_inputs,
                                                    get_input_bits(plc_id), sampling_frequency)
                        register_task = read_registers(client, client.read_input_registers,
                                                       get_inputs_register(plc_id), sampling_frequency, 125)

                        coil_states, input_states, register_states = await asyncio.gather(
                            coil_task, input_task, register_task
                        )

                        # Thread-safe data update
                        with data_lock:
                            latest_modbus_data[plc_id] = {
                                "coil_states": coil_states ,
                                "input_status_states": input_states ,
                                "input_register_states": register_states
                            }
                            logger.info(f"latest_modbus_data : {latest_modbus_data}")
                        # Save to Excel
                        await append_to_excel({
                            "plc_id": plc_id,
                            "coil_states": coil_states,
                            "input_status_states": input_states,
                            "input_register_states": register_states
                        })

                    except ModbusException as e:
                        logger.error(f"Modbus error in {plc_id}: {e}")
                        await asyncio.sleep(1)
                    except Exception as e:
                        logger.error(f"Unexpected error in {plc_id}: {e}")
                        await asyncio.sleep(1)

        except ConnectionException as e:
            logger.error(f"Connection failed for {plc_id}: {e}")
            await asyncio.sleep(5)
        except Exception as e:
            logger.error(f"Critical error for {plc_id}: {e}")
            await asyncio.sleep(10)

