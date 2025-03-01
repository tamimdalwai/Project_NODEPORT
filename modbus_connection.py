import asyncio
from pymodbus import ModbusException
from pymodbus.client import AsyncModbusTcpClient
from pymodbus.exceptions import ConnectionException
from tkinter import Tk, messagebox
import socket
import os
from read_plc import get_coils, get_input_bits, get_analog_inputs, read_selected_plc
from openpyxl import Workbook, load_workbook
from datetime import datetime
from openpyxl.utils.exceptions import InvalidFileException
from zipfile import BadZipFile
import logging
from map import map_value

# from home_ui import HomeUI

# from home_ui import HomeUI  # Importing the HomeUI class

#     # # Initialize Tkinter window and HomeUI instance
# import tkinter as tk
# root = tk.Tk()

#     # Create the HomeUI instance and pack it
# home_ui_instance = HomeUI(root)
# home_ui_instance.pack()

# Getting selected PLC and address info
selected_plc = read_selected_plc()

print(f"Selected PLC: {selected_plc}")

# Global variable to track connection states
connected = False

# Path to the Excel file
EXCEL_FILE_PATH = r"output\analog_register_data.xlsx"
saved_address_file_path = r"config\saveAddress.xlsx"

coils_count = get_coils(saved_address_file_path, selected_plc)
print(f"Coils count: {coils_count}")

input_bits_count = get_input_bits(saved_address_file_path, selected_plc)
print(f"Input bits count: {input_bits_count}")

input_registers_count = get_analog_inputs(saved_address_file_path, selected_plc)
print(f"Input registers count: {input_registers_count}")

_clist1 = None


# Initialize PLC state dictionary to track connections and previous data
def init_plc_state(plc_id):
    plc_state[plc_id] = {
        "connected": False,
        "previous_input_registers": None
    }


# plc_state = {selected_plc: {"connected": False, "previous_input_registers": None}}
plc_state = {selected_plc: {"connected": False, "previous_input_registers": None}}


# Check if the data has changed since the last read
def has_data_changed(plc_id, current_data, data_type):
    previous_data = plc_state[plc_id].get(f"previous_{data_type}")
    if previous_data is None or current_data != previous_data:
        plc_state[plc_id][f"previous_{data_type}"] = current_data
        return True
    return False



def write_to_excel(plc_name, input_registers, ip_address, port_number):
    """
    Writes data to an Excel file, ensuring values are mapped between 0 and 65000.
    """

    # Ensure input_registers is a list
    if not isinstance(input_registers, list):
        print(f"Expected a list for input_registers, but got {type(input_registers)}. Converting...")
        input_registers = [input_registers] if input_registers is not None else []

    # Convert values to floats and ensure they fall within 0 - 65000
    try:
        input_registers = [
            max(0, min(65000, float(value))) if value is not None else None
            for value in input_registers
        ]
    except ValueError as e:
        print(f"Error converting register values to float: {e}")
        input_registers = []

    # Ensure there are exactly 128 register values (Pad with None if necessary)
    input_registers += [None] * (128 - len(input_registers))

    # Get current timestamp
    current_time = datetime.now()
    date = current_time.strftime("%Y-%m-%d")
    time = current_time.strftime("%H:%M")

    # Data to be written
    data = [date, time, ip_address, port_number] + input_registers[:128]

    try:
        # Load existing workbook if it exists, otherwise create a new one
        if os.path.exists(EXCEL_FILE_PATH):
            try:
                wb = load_workbook(EXCEL_FILE_PATH)
                if plc_name not in wb.sheetnames:
                    ws = wb.create_sheet(plc_name)
                    headers = ["Date", "Time", "IP Address", "Port Number"] + [f"{300001 + i}" for i in range(128)]
                    ws.append(headers)
                else:
                    ws = wb[plc_name]
            except (InvalidFileException, BadZipFile):
                print(f"File {EXCEL_FILE_PATH} is corrupted. Recreating file.")
                os.remove(EXCEL_FILE_PATH)
                wb = Workbook()
                ws = wb.active
                ws.title = plc_name
                headers = ["Date", "Time", "IP Address", "Port Number"] + [f"{300001 + i}" for i in range(128)]
                ws.append(headers)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = plc_name
            headers = ["Date", "Time", "IP Address", "Port Number"] + [f"{300001 + i}" for i in range(128)]
            ws.append(headers)

        # Append data while preserving decimal values
        for i in range(4, len(data)):  # Skip first 4 columns (Date, Time, IP, Port)
            if isinstance(data[i], (int, float)):
                data[i] = round(float(data[i]), 4)  # Keep 4 decimal places

        ws.append(data)
        wb.save(EXCEL_FILE_PATH)
        print(f"Data written to sheet '{plc_name}' in Excel file.")
    except PermissionError:
        print("Permission denied: The Excel file is currently open. Please close it to write data.")
    except Exception as e:
        print(f"An error occurred while writing to the Excel file: {e}")

async def async_write_to_excel(data, plc_id):
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, write_to_excel, *data)


# Modbus client functions for reading coils, input registers, and input statuses
async def run_modbus_client_coil(client, sampling_frequency, plc_id):
    try:

        coil_result = await client.read_coils(0, coils_count)
        if coil_result.isError():
            print(f"Error reading coils for PLC {plc_id}: {coil_result}")
            # await asyncio.sleep(5)
            return None
        await asyncio.sleep(sampling_frequency / 1000)
        print(coil_result.bits)
        global _clist1
        if _clist1 is None:
            _clist1 = coil_result.bits
        return coil_result.bits
    except Exception as e:
        print(f"PLC {plc_id} error reading coils: {e}")
        await asyncio.sleep(5)
        # await check_connection(client, plc_id)
        return None


async def run_modbus_client_input_status(client, sampling_frequency, plc_id):
    try:
        input_status_result = await client.read_discrete_inputs(0, input_bits_count)
        if input_status_result.isError():
            # await asyncio.sleep(5)
            return None
        await asyncio.sleep(sampling_frequency / 1000)
        print(input_status_result.bits)
        return input_status_result.bits
    except Exception as e:
        print(f"PLC {plc_id} error reading input statuses: {e}")
        # await asyncio.sleep(5)
        # await check_connection(client, plc_id)
        return None


async def run_modbus_client_input_register(client, sampling_frequency, plc_id):
    try:
        print("in run_modbus_client_input_register function ")
        print("input_registers_count",input_registers_count)
        input_register_result = await client.read_input_registers(0, input_registers_count)
        if input_register_result.isError():
            print(f"PLC {plc_id} error: Input register read failed.")
            # await asyncio.sleep(5)
            return None

        await asyncio.sleep(sampling_frequency / 1000)

        # Ensure input_register_result contains registers
        if not hasattr(input_register_result, "registers") or not input_register_result.registers:
            print(f"PLC {plc_id} error: No registers found in input register result.")
            return None

        current_input_registers = input_register_result.registers
        # print("input register: ",input_register_result.registers)
        print(f"PLC {plc_id} mapped value: {current_input_registers}")
        return current_input_registers

    except (Exception, ConnectionException, socket.error, ModbusException) as e:
        print(f"PLC {plc_id} error reading input registers: {e}")
        user_response = show_popup_and_wait_input_register(f"Connection failed: {e}. Click Retry to check again.")
        if not user_response:
            print("User canceled retry. Stopping connection attempts.")
            exit(0)
        # await check_connection(client, plc_id)

        return None


def show_popup_and_wait_input_register(message):
    """
    Displays a Tkinter popup with a Retry/Cancel option.
    Returns True for Retry, False for Cancel.
    """
    import tkinter as tk
    root = tk.Tk()  # type: ignore
    root.withdraw()
    response = messagebox.askretrycancel("Connection Error, Click On Retry", message)
    root.destroy()
    return response


def show_popup_and_wait(message):
    """
    Displays a Tkinter popup with a Retry/Cancel option.
    Returns True for Retry, False for Cancel.
    """
    import tkinter as tk
    root = tk.Tk()  # type: ignore
    root.withdraw()
    messagebox.showwarning("Connection Error,Start PLC", message)
    root.destroy()


async def check_connection(client, plc_id, retry_interval=5):
    try:
        print("come in check")
        # await asyncio.sleep(5)
        input_register_check = await client.read_input_registers(0, input_bits_count)
        # await asyncio.sleep(1)
        input_register_ok = not input_register_check.isError()
        print("after input_register check")
        print(coils_count)
        coil_check = await client.read_coils(0, coils_count)
        print("in coil check")
        coil_ok = not coil_check.isError()
        print("after coil check")

        input_status_check = await client.read_discrete_inputs(0, input_registers_count)
        input_status_ok = not input_status_check.isError()

        plc_state[plc_id]["connected"] = coil_ok or input_register_ok or input_status_ok
        print(f"PLC {plc_id} connected")
        return plc_state[plc_id]["connected"]

    except (ConnectionException, socket.error, ModbusException, Exception) as e:
        print(f"{plc_id} connection error from check_connection: {e}")
        plc_state[plc_id]["connected"] = False
        print("above popup")
        # await asyncio.sleep(1)
        user_response = show_popup_and_wait(f"Connection failed: {e}. Click Retry to check again.")
        if not user_response:
            print("User canceled retry. Stopping connection attempts.")
            exit(0)
        print("below pop up")
        # connect_retry(client, plc_id,retry_interval=5)
    await asyncio.sleep(retry_interval)
    print("going up")



# Main Modbus client loop for a specific PLC
async def modbus_client_loop(ip_address, port, sampling_frequency, update_callback, plc_id):  #,home_ui_instance
    init_plc_state(plc_id)  # Initialize the PLC state
    try:
        async with AsyncModbusTcpClient(ip_address, port=port) as client:
            print("u ")
            while True:
                print("in while loop of modbus loop ")
                # Check connection every loop iteration

                if not plc_state[plc_id]["connected"]:
                    if await check_connection(client, plc_id):
                        print(f"PLC {plc_id} connected. Starting data retrieval...")
                    else:
                        print(f"PLC {plc_id} not connected. Retrying...")
                        await check_connection(client, plc_id)
                # input_register_result = await client.read_input_registers(0, input_registers_count)
                # print(input_register_result)
                # Run Modbus client methods in parallel
                print("above gather")
                input_registers, coil_states, input_statuses = await asyncio.gather(
                    run_modbus_client_input_register(client, sampling_frequency, plc_id),
                    run_modbus_client_coil(client, sampling_frequency, plc_id),
                    run_modbus_client_input_status(client, sampling_frequency, plc_id)
                )
                print("below gather")
                # Handle None values
                coil_states = coil_states or []
                input_registers = input_registers or []
                input_statuses = input_statuses or []
                # Call update_ui_wrapper method of HomeUI to update UI
                # home_ui_instance.update_ui_wrapper(coil_states, input_statuses, input_registers,selected_plc)
                # Call the update callback instead of directly updating the UI
                # selected_plc = HomeUI.getSelectedPLC();
                selected_plc = read_selected_plc()
                if update_callback:
                    if selected_plc == plc_id:
                        print("in update_callback")
                        print(selected_plc, "and", plc_id)
                        update_callback(coil_states, input_statuses, input_registers, plc_id)
                print("out of update callback")
                # If data has been read successfully and has changed, write to Excel
                if input_registers and has_data_changed(plc_id, input_registers, "input_registers"):
                    try:
                        await async_write_to_excel(
                            (plc_id, input_registers, ip_address, port),
                            plc_id
                        )
                        print(f"Data written to sheet for PLC {plc_id} in Excel file.")
                    except Exception as e:
                        print(f"Error writing data for PLC {plc_id}: {e}")

                # # Handle retries if there are issues with reading coils, input bits, or registers
                # if not coil_states or not input_registers or not input_statuses:
                #     print(f"Retrying data read for PLC {plc_id} due to missing data.")
                #     await asyncio.sleep(5)  # Sleep before retrying
                #     continue
                await asyncio.sleep(sampling_frequency / 1000)
                print(selected_plc, "and", plc_id)
    except (ConnectionException, ModbusException, Exception) as e:
        print(f"from modbus_clinet_loop Connection failed for PLC {plc_id}: {e}")
        await check_connection(client, plc_id)
# async def update_callback(coil_states, input_statuses, input_registers, plc_id):
#     print(f"PLC {plc_id} Data: Coils={coil_states}, Input Statuses={input_statuses}, Input Registers={input_registers}")
# async def main():
#     await modbus_client_loop("192.168.1.7", 502, 1000, update_callback, "PLC1")
#
# asyncio.run(main())
